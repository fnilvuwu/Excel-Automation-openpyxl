from openpyxl import load_workbook

# Load both workbooks
wb1 = load_workbook("Training History.xlsx")
wb2 = load_workbook("Employee.xlsx")

# Access the active sheet for each workbook
ws1 = wb1.active
ws2 = wb2.active

# Create a dictionary to store employee_id to division_name mappings
division_names = {}

# Iterate through all rows in ws2 to build the dictionary
for row_index in range(1, ws2.max_row + 1):
    employee_id = ws2.cell(row=row_index, column=1).value  # Accessing column A
    division_name = ws2.cell(row=row_index, column=7).value  # Accessing column G
    division_names[employee_id] = division_name

# Iterate through rows 4 to 107082 in ws1
for row in range(4, 107083):  # Assuming there are 107082 rows to check
    employee_id = ws1["B" + str(row)].value
    print(f"Searching for Employee ID {employee_id} in ws2...")
    
    # Check if employee_id is None
    if employee_id is not None:
        # Check if employee_id exists in division_names dictionary
        if employee_id in division_names:
            # Retrieve division_name from the dictionary
            division_name = division_names[employee_id]
            ws1["D" + str(row)] = division_name
            print(f"Division name for Employee ID {employee_id}: {division_name}")
        else:
            print(f"Employee ID {employee_id} not found in ws2.")
    else:
        print("Employee ID is None, skipping.")

wb1.save("Edited_Training History.xlsx")
